from __future__ import annotations

import argparse
import hashlib
import html
import json
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import date, datetime, time
from datetime import timedelta
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_HEADERS = [
    "出单员",
    "退回审核意见",
    "提核退回标志",
    "出单时间",
    "保单号",
    "投保单号",
]
OPERATOR_PATTERN = re.compile(r"([\u3400-\u9fff·]{2,12})$")


def text_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value)


def parse_record_date(value: object) -> tuple[str | None, int | None, int | None]:
    if value is None or text_value(value).strip() == "":
        return None, None, None
    if isinstance(value, datetime):
        return value.date().isoformat(), value.year, value.month
    if isinstance(value, date):
        return value.isoformat(), value.year, value.month

    value_text = text_value(value).strip()
    match = re.match(r"^(20\d{2})[-/.](\d{1,2})[-/.](\d{1,2})", value_text)
    if not match:
        return None, None, None
    year, month, day = (int(part) for part in match.groups())
    try:
        parsed = date(year, month, day)
    except ValueError:
        return None, None, None
    return parsed.isoformat(), parsed.year, parsed.month


def operator_name(value: str) -> str:
    normalized = value.strip()
    for prefix in ("客户经理", "业务员", "出单员", "操作员"):
        if normalized.startswith(prefix) and len(normalized) > len(prefix):
            normalized = normalized[len(prefix):]
            break
    match = OPERATOR_PATTERN.search(normalized)
    return match.group(1) if match else ""


def row_hash(values: dict[str, str]) -> str:
    canonical_values = {header: values[header] for header in sorted(values)}
    payload = json.dumps(canonical_values, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def excel_column_number(reference: str) -> int:
    letters = re.match(r"[A-Z]+", reference.upper()).group(0)
    number = 0
    for letter in letters:
        number = number * 26 + ord(letter) - ord("A") + 1
    return number


def iter_xml_blocks(stream, start_token: bytes, end_token: bytes):
    buffer = b""
    chunk_size = 8 * 1024 * 1024
    while True:
        chunk = stream.read(chunk_size)
        if chunk:
            buffer += chunk
        while True:
            start = buffer.find(start_token)
            if start < 0:
                buffer = buffer[-len(start_token):]
                break
            end = buffer.find(end_token, start)
            if end < 0:
                buffer = buffer[start:]
                break
            end += len(end_token)
            yield buffer[start:end]
            buffer = buffer[end:]
        if not chunk:
            break


def fast_cell_value(cell: bytes, shared_strings: list[str]) -> tuple[int, str]:
    ref_match = re.search(rb'\br="([A-Z]+)\d+"', cell)
    if not ref_match:
        return 0, ""
    column = excel_column_number(ref_match.group(1).decode("ascii"))
    type_match = re.search(rb'\bt="([^"]+)"', cell)
    value_match = re.search(rb"<v>(.*?)</v>", cell, flags=re.S)
    raw = value_match.group(1).decode("utf-8", "replace") if value_match else ""
    if type_match and type_match.group(1) == b"s" and raw:
        raw = shared_strings[int(raw)]
    elif type_match and type_match.group(1) == b"inlineStr":
        raw = "".join(match.decode("utf-8", "replace") for match in re.findall(rb"<t[^>]*>(.*?)</t>", cell, flags=re.S))
    return column, html.unescape(raw)


def excel_column_letters(number: int) -> bytes:
    letters = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters.encode("ascii")


def xml_cell_raw(cell) -> str:
    value = cell.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
    return "" if value is None else value.text or ""


def read_selected_shared_strings(archive, wanted: set[int]) -> dict[int, str]:
    values = {}
    if not wanted:
        return values
    with archive.open("xl/sharedStrings.xml") as stream:
        index = 0
        for _, item in ET.iterparse(stream, events=("end",)):
            if item.tag == "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si":
                if index in wanted:
                    values[index] = "".join(item.itertext())
                item.clear()
                index += 1
    return values


def parse_large_workbook_selected(source: Path) -> dict[str, object]:
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with zipfile.ZipFile(source) as archive:
        header_raw = {}
        with archive.open("xl/worksheets/sheet1.xml") as stream:
            for _, row in ET.iterparse(stream, events=("end",)):
                if row.tag != namespace + "row":
                    continue
                if row.attrib.get("r") == "1":
                    for cell in row.findall(namespace + "c"):
                        header_raw[re.match(r"[A-Z]+", cell.attrib.get("r", "")).group(0)] = (cell.attrib.get("t", ""), xml_cell_raw(cell))
                    row.clear()
                    break
                row.clear()
        header_indices = {column: int(raw) for column, (kind, raw) in header_raw.items() if kind == "s" and raw.isdigit()}
        header_strings = read_selected_shared_strings(archive, set(header_indices.values()))
        headers_by_column = {column: header_strings.get(index, "") for column, index in header_indices.items()}
        selected_names = ["出单员", "退回审核意见", "提核退回标志", "出单时间", "保单号", "投保单号"]
        field_cols = {name: column for column, name in headers_by_column.items()}
        missing = [name for name in selected_names if name not in field_cols]
        if missing:
            return {"sourceFile": str(source), "headers": list(headers_by_column.values()), "rows": [], "periods": [], "errors": [{"code": "missing_headers", "headers": missing, "message": f"缺少必需表头: {'、'.join(missing)}"}]}

        selected_cols = set(field_cols.values())
        records = []
        wanted_indices: set[int] = set()
        with archive.open("xl/worksheets/sheet1.xml") as stream:
            for _, row in ET.iterparse(stream, events=("end",)):
                if row.tag != namespace + "row":
                    continue
                row_number = int(row.attrib.get("r", "0"))
                if row_number > 1:
                    row_cells = {}
                    for cell in row.findall(namespace + "c"):
                        column = re.match(r"[A-Z]+", cell.attrib.get("r", "")).group(0)
                        if column not in selected_cols:
                            continue
                        kind, raw = cell.attrib.get("t", ""), xml_cell_raw(cell)
                        row_cells[column] = (kind, raw)
                        if kind == "s" and raw.isdigit():
                            wanted_indices.add(int(raw))
                    records.append((row_number, row_cells))
                row.clear()
        strings = read_selected_shared_strings(archive, wanted_indices)

        def value(row_cells, name):
            kind, raw = row_cells.get(field_cols[name], ("", ""))
            return strings.get(int(raw), "") if kind == "s" and raw.isdigit() else raw

        parsed_rows = []
        errors = []
        periods: set[tuple[int, int]] = set()
        for source_row, row_cells in records:
            opinion = value(row_cells, "退回审核意见").strip()
            if not opinion:
                continue
            values = {name: value(row_cells, name) for name in selected_names}
            date_value = values["出单时间"]
            if re.match(r"^\d+(\.\d+)?$", date_value):
                date_value = (date(1899, 12, 30) + timedelta(days=float(date_value))).isoformat()
            record_date, year, month = parse_record_date(date_value)
            if record_date is None:
                errors.append({"code": "missing_date", "sourceRow": source_row, "header": "出单时间", "message": f"第 {source_row} 行缺少有效的出单时间"})
            else:
                periods.add((year, month))
            normalized = {
                "sourceRow": source_row, "values": values, "operatorRaw": values["出单员"],
                "operatorName": operator_name(values["出单员"]), "opinion": opinion,
                "returnFlag": values["提核退回标志"], "recordDate": record_date, "year": year, "month": month,
                "rowKey": values["保单号"].strip() or values["投保单号"].strip(),
            }
            normalized["rowHash"] = row_hash(values)
            parsed_rows.append(normalized)
    return {"sourceFile": str(source), "headers": selected_names, "rows": parsed_rows, "periods": [{"year": year, "month": month} for year, month in sorted(periods)], "errors": errors}


def parse_large_workbook_targeted(source: Path) -> dict[str, object]:
    with zipfile.ZipFile(source) as archive:
        shared_strings = []
        if "xl/sharedStrings.xml" in archive.namelist():
            with archive.open("xl/sharedStrings.xml") as shared_stream:
                for item in iter_xml_blocks(shared_stream, b"<si", b"</si>"):
                    shared_strings.append(html.unescape(b"".join(re.findall(rb"<t[^>]*>(.*?)</t>", item, flags=re.S)).decode("utf-8", "replace")))
        sheet_data = archive.read("xl/worksheets/sheet1.xml")

    row_pattern = re.compile(rb"<row\b[^>]*\br=\"(\d+)\"[^>]*>.*?</row>", flags=re.S)
    first_row = row_pattern.search(sheet_data)
    if not first_row:
        return {"sourceFile": str(source), "headers": [], "rows": [], "periods": [], "errors": [{"code": "empty_sheet", "message": "工作表没有数据"}]}
    first_cells = {
        column: value for column, value in (
            fast_cell_value(cell, shared_strings)
            for cell in re.findall(rb"<c\b.*?</c>", first_row.group(0), flags=re.S)
        ) if column
    }
    headers = [first_cells.get(column, "").strip() for column in range(1, max(first_cells, default=0) + 1)]
    header_columns = {header: index for index, header in enumerate(headers, start=1)}
    missing_headers = [header for header in REQUIRED_HEADERS if header not in header_columns]
    if missing_headers:
        return {"sourceFile": str(source), "headers": headers, "rows": [], "periods": [], "errors": [{"code": "missing_headers", "headers": missing_headers, "message": f"缺少必需表头: {'、'.join(missing_headers)}"}]}

    opinion_column = excel_column_letters(header_columns["退回审核意见"])
    opinion_pattern = re.compile(rb"<c\b[^>]*\br=\"" + opinion_column + rb"(\d+)\"[^>]*>.*?</c>", flags=re.S)
    candidate_rows = set()
    for match in opinion_pattern.finditer(sheet_data):
        _, opinion = fast_cell_value(match.group(0), shared_strings)
        if opinion.strip():
            candidate_rows.add(int(match.group(1)))

    parsed_rows: list[dict[str, object]] = []
    errors: list[dict[str, object]] = []
    periods: set[tuple[int, int]] = set()
    cells_pattern = re.compile(rb"<c\b.*?</c>", flags=re.S)
    for row_match in row_pattern.finditer(sheet_data):
        source_row = int(row_match.group(1))
        if source_row not in candidate_rows:
            continue
        cells = {column: value for column, value in (fast_cell_value(cell, shared_strings) for cell in cells_pattern.findall(row_match.group(0))) if column}
        values = {header: cells.get(index, "") for index, header in enumerate(headers, start=1)}
        date_value = values["出单时间"]
        if re.match(r"^\d+(\.\d+)?$", date_value):
            date_value = (date(1899, 12, 30) + timedelta(days=float(date_value))).isoformat()
        record_date, year, month = parse_record_date(date_value)
        if record_date is None:
            errors.append({"code": "missing_date", "sourceRow": source_row, "header": "出单时间", "message": f"第 {source_row} 行缺少有效的出单时间"})
        else:
            periods.add((year, month))
        normalized = {
            "sourceRow": source_row, "values": values, "operatorRaw": values["出单员"],
            "operatorName": operator_name(values["出单员"]), "opinion": values["退回审核意见"],
            "returnFlag": values["提核退回标志"], "recordDate": record_date, "year": year, "month": month,
            "rowKey": values["保单号"].strip() or values["投保单号"].strip(),
        }
        normalized["rowHash"] = row_hash(values)
        parsed_rows.append(normalized)
    return {"sourceFile": str(source), "headers": headers, "rows": parsed_rows, "periods": [{"year": year, "month": month} for year, month in sorted(periods)], "errors": errors}


def parse_large_workbook_fast(source: Path) -> dict[str, object]:
    with zipfile.ZipFile(source) as archive:
        shared_strings = []
        if "xl/sharedStrings.xml" in archive.namelist():
            with archive.open("xl/sharedStrings.xml") as shared_stream:
                for item in iter_xml_blocks(shared_stream, b"<si", b"</si>"):
                    shared_strings.append(html.unescape(b"".join(re.findall(rb"<t[^>]*>(.*?)</t>", item, flags=re.S)).decode("utf-8", "replace")))

        headers: list[str] = []
        parsed_rows: list[dict[str, object]] = []
        errors: list[dict[str, object]] = []
        periods: set[tuple[int, int]] = set()
        header_columns: dict[str, int] = {}
        with archive.open("xl/worksheets/sheet1.xml") as sheet_stream:
            for row_block in iter_xml_blocks(sheet_stream, b"<row", b"</row>"):
                cells = {column: value for column, value in (
                    fast_cell_value(cell, shared_strings)
                    for cell in re.findall(rb"<c\b.*?</c>", row_block, flags=re.S)
                ) if column}
                if not headers:
                    max_column = max(cells, default=0)
                    headers = [cells.get(column, "").strip() for column in range(1, max_column + 1)]
                    header_columns = {header: index for index, header in enumerate(headers, start=1)}
                    missing_headers = [header for header in REQUIRED_HEADERS if header not in header_columns]
                    if missing_headers:
                        return {
                            "sourceFile": str(source), "headers": headers, "rows": [], "periods": [],
                            "errors": [{"code": "missing_headers", "headers": missing_headers, "message": f"缺少必需表头: {'、'.join(missing_headers)}"}],
                        }
                    continue

                opinion = cells.get(header_columns["退回审核意见"], "").strip()
                if not opinion:
                    continue
                values = {header: cells.get(index, "") for index, header in enumerate(headers, start=1)}
                date_value = values["出单时间"]
                if re.match(r"^\d+(\.\d+)?$", date_value):
                    date_value = (date(1899, 12, 30) + timedelta(days=float(date_value))).isoformat()
                record_date, year, month = parse_record_date(date_value)
                row_match = re.search(rb'\br="(\d+)"', row_block)
                source_row = int(row_match.group(1)) if row_match else len(parsed_rows) + 2
                if record_date is None:
                    errors.append({"code": "missing_date", "sourceRow": source_row, "header": "出单时间", "message": f"第 {source_row} 行缺少有效的出单时间"})
                else:
                    periods.add((year, month))
                operator_raw = values["出单员"]
                normalized = {
                    "sourceRow": source_row, "values": values, "operatorRaw": operator_raw,
                    "operatorName": operator_name(operator_raw), "opinion": opinion,
                    "returnFlag": values["提核退回标志"], "recordDate": record_date,
                    "year": year, "month": month,
                    "rowKey": values["保单号"].strip() or values["投保单号"].strip(),
                }
                normalized["rowHash"] = row_hash(values)
                parsed_rows.append(normalized)
    return {"sourceFile": str(source), "headers": headers, "rows": parsed_rows, "periods": [{"year": year, "month": month} for year, month in sorted(periods)], "errors": errors}


def parse_large_workbook(source: Path) -> dict[str, object]:
    from lxml import etree

    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    shared_strings: list[str] = []
    with zipfile.ZipFile(source) as archive:
        if "xl/sharedStrings.xml" in archive.namelist():
            with archive.open("xl/sharedStrings.xml") as shared_stream:
                for _, item in etree.iterparse(shared_stream, events=("end",), tag=f"{namespace}si"):
                    shared_strings.append("".join(item.itertext()))
                    item.clear()

        with archive.open("xl/worksheets/sheet1.xml") as sheet_stream:
            rows = etree.iterparse(sheet_stream, events=("end",), tag=f"{namespace}row")
            headers: list[str] = []
            parsed_rows: list[dict[str, object]] = []
            errors: list[dict[str, object]] = []
            periods: set[tuple[int, int]] = set()
            header_columns: dict[str, int] = {}
            for _, row_element in rows:
                cell_values: dict[int, str] = {}
                for cell in row_element:
                    if cell.tag != f"{namespace}c":
                        continue
                    reference = cell.attrib.get("r", "")
                    column = excel_column_number(reference)
                    value_node = cell.find(f"{namespace}v")
                    inline_text = cell.find(f"{namespace}is")
                    raw = "" if value_node is None else value_node.text or ""
                    if inline_text is not None:
                        raw = "".join(inline_text.itertext())
                    if cell.attrib.get("t") == "s" and raw:
                        raw = shared_strings[int(raw)]
                    cell_values[column] = raw

                if not headers:
                    max_column = max(cell_values, default=0)
                    headers = [cell_values.get(column, "").strip() for column in range(1, max_column + 1)]
                    header_columns = {header: index for index, header in enumerate(headers, start=1)}
                    missing_headers = [header for header in REQUIRED_HEADERS if header not in header_columns]
                    if missing_headers:
                        return {
                            "sourceFile": str(source),
                            "headers": headers,
                            "rows": [],
                            "periods": [],
                            "errors": [{
                                "code": "missing_headers",
                                "headers": missing_headers,
                                "message": f"缺少必需表头: {'、'.join(missing_headers)}",
                            }],
                        }
                    row_element.clear()
                    continue

                opinion = text_value(cell_values.get(header_columns["退回审核意见"], "")).strip()
                if not opinion:
                    row_element.clear()
                    continue
                values = {
                    header: text_value(cell_values.get(index, ""))
                    for index, header in enumerate(headers, start=1)
                }
                date_value = cell_values.get(header_columns["出单时间"], "")
                if re.match(r"^\d+(\.\d+)?$", str(date_value)):
                    date_value = (date(1899, 12, 30) + timedelta(days=float(date_value))).isoformat()
                record_date, year, month = parse_record_date(date_value)
                source_row = int(row_element.attrib.get("r", len(parsed_rows) + 2))
                if record_date is None:
                    errors.append({
                        "code": "missing_date",
                        "sourceRow": source_row,
                        "header": "出单时间",
                        "message": f"第 {source_row} 行缺少有效的出单时间",
                    })
                else:
                    periods.add((year, month))
                operator_raw = values["出单员"]
                policy_no = values["保单号"].strip()
                proposal_no = values["投保单号"].strip()
                normalized = {
                    "sourceRow": source_row,
                    "values": values,
                    "operatorRaw": operator_raw,
                    "operatorName": operator_name(operator_raw),
                    "opinion": opinion,
                    "returnFlag": values["提核退回标志"],
                    "recordDate": record_date,
                    "year": year,
                    "month": month,
                    "rowKey": policy_no or proposal_no,
                }
                normalized["rowHash"] = row_hash(values)
                parsed_rows.append(normalized)
                row_element.clear()

    return {
        "sourceFile": str(source),
        "headers": headers,
        "rows": parsed_rows,
        "periods": [{"year": year, "month": month} for year, month in sorted(periods)],
        "errors": errors,
    }


def parse_workbook(source: Path) -> dict[str, object]:
    if source.stat().st_size >= 50 * 1024 * 1024:
        return parse_large_workbook_selected(source)
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header_values = next((list(row) for row in rows if any(value is not None for value in row)), [])
        headers = [text_value(value).strip() for value in header_values]
        missing_headers = [header for header in REQUIRED_HEADERS if header not in headers]
        if missing_headers:
            return {
                "sourceFile": str(source),
                "headers": headers,
                "rows": [],
                "periods": [],
                "errors": [{
                    "code": "missing_headers",
                    "headers": missing_headers,
                    "message": f"缺少必需表头: {'、'.join(missing_headers)}",
                }],
            }

        header_index = {header: index for index, header in enumerate(headers)}
        parsed_rows: list[dict[str, object]] = []
        errors: list[dict[str, object]] = []
        periods: set[tuple[int, int]] = set()
        for source_row, raw_row in enumerate(rows, start=2):
            values = {
                header: text_value(raw_row[index] if index < len(raw_row) else None)
                for index, header in enumerate(headers)
            }
            if not any(values.values()):
                continue
            if not values["退回审核意见"].strip():
                continue

            record_date, year, month = parse_record_date(raw_row[header_index["出单时间"]])
            if record_date is None:
                errors.append({
                    "code": "missing_date",
                    "sourceRow": source_row,
                    "header": "出单时间",
                    "message": f"第 {source_row} 行缺少有效的出单时间",
                })
            else:
                periods.add((year, month))

            operator_raw = values["出单员"]
            policy_no = values["保单号"].strip()
            proposal_no = values["投保单号"].strip()
            normalized = {
                "sourceRow": source_row,
                "values": values,
                "operatorRaw": operator_raw,
                "operatorName": operator_name(operator_raw),
                "opinion": values["退回审核意见"],
                "returnFlag": values["提核退回标志"],
                "recordDate": record_date,
                "year": year,
                "month": month,
                "rowKey": policy_no or proposal_no,
            }
            normalized["rowHash"] = row_hash(values)
            parsed_rows.append(normalized)

        return {
            "sourceFile": str(source),
            "headers": headers,
            "rows": parsed_rows,
            "periods": [
                {"year": year, "month": month}
                for year, month in sorted(periods)
            ],
            "errors": errors,
        }
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    result = parse_workbook(args.input.resolve())
    args.output.resolve().parent.mkdir(parents=True, exist_ok=True)
    args.output.resolve().write_text(
        json.dumps(result, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
