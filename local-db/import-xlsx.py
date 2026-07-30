from __future__ import annotations

import hashlib
import json
import re
import sqlite3
import sys
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


DB_PATH = Path(__file__).with_name("work.db")
MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


SCHEMA = """
CREATE TABLE IF NOT EXISTS spreadsheet_workbooks (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  record_id INTEGER NOT NULL UNIQUE,
  attachment_id INTEGER NOT NULL UNIQUE,
  filename TEXT NOT NULL,
  report_year INTEGER,
  report_month INTEGER,
  document_type TEXT NOT NULL,
  sheet_count INTEGER NOT NULL,
  populated_cell_count INTEGER NOT NULL DEFAULT 0,
  formula_count INTEGER NOT NULL DEFAULT 0,
  imported_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
  FOREIGN KEY (record_id) REFERENCES records(id) ON DELETE CASCADE,
  FOREIGN KEY (attachment_id) REFERENCES attachments(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_spreadsheet_workbooks_period
  ON spreadsheet_workbooks(report_year, report_month);
CREATE INDEX IF NOT EXISTS idx_spreadsheet_workbooks_type
  ON spreadsheet_workbooks(document_type);

CREATE TABLE IF NOT EXISTS spreadsheet_sheets (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  workbook_id INTEGER NOT NULL,
  sheet_index INTEGER NOT NULL,
  sheet_name TEXT NOT NULL,
  max_row INTEGER NOT NULL,
  max_column INTEGER NOT NULL,
  sheet_state TEXT NOT NULL,
  merged_ranges_json TEXT NOT NULL DEFAULT '[]',
  FOREIGN KEY (workbook_id) REFERENCES spreadsheet_workbooks(id) ON DELETE CASCADE,
  UNIQUE (workbook_id, sheet_index)
);
CREATE INDEX IF NOT EXISTS idx_spreadsheet_sheets_workbook
  ON spreadsheet_sheets(workbook_id);

CREATE TABLE IF NOT EXISTS spreadsheet_cells (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  sheet_id INTEGER NOT NULL,
  row_number INTEGER NOT NULL,
  column_number INTEGER NOT NULL,
  cell_address TEXT NOT NULL,
  value_type TEXT NOT NULL,
  value_text TEXT,
  formula_text TEXT,
  cached_value_text TEXT,
  number_format TEXT,
  FOREIGN KEY (sheet_id) REFERENCES spreadsheet_sheets(id) ON DELETE CASCADE,
  UNIQUE (sheet_id, row_number, column_number)
);
CREATE INDEX IF NOT EXISTS idx_spreadsheet_cells_sheet_position
  ON spreadsheet_cells(sheet_id, row_number, column_number);
"""


def text_value(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def classify(filename: str) -> tuple[int | None, int | None, str]:
    year_match = re.search(r"(20\d{2})", filename)
    month_match = re.search(r"(?:年|\.)(1[0-2]|[1-9])月?", filename)
    year = int(year_match.group(1)) if year_match else None
    month = int(month_match.group(1)) if month_match else None
    if "考核" in filename:
        document_type = "assessment"
    elif "报人力" in filename:
        document_type = "hr_piecework_payroll"
    elif "计件工资" in filename:
        document_type = "piecework_payroll"
    else:
        document_type = "spreadsheet"
    return year, month, document_type


def import_workbook(connection: sqlite3.Connection, source: Path) -> str:
    content = source.read_bytes()
    sha256 = hashlib.sha256(content).hexdigest()
    if connection.execute(
        "SELECT 1 FROM attachments WHERE sha256 = ?", (sha256,)
    ).fetchone():
        return f"Skipped duplicate: {source.name}"

    formula_book = load_workbook(source, data_only=False, read_only=False)
    value_book = load_workbook(source, data_only=True, read_only=False)
    year, month, document_type = classify(source.name)
    metadata = {
        "filename": source.name,
        "source_path": str(source),
        "size_bytes": len(content),
        "sha256": sha256,
        "file_type": "xlsx",
        "report_year": year,
        "report_month": month,
        "document_type": document_type,
    }

    with connection:
        record_cursor = connection.execute(
            "INSERT INTO records (category, title, data_json) VALUES (?, ?, ?)",
            ("spreadsheet", source.name, json.dumps(metadata, ensure_ascii=False)),
        )
        record_id = record_cursor.lastrowid
        attachment_cursor = connection.execute(
            """
            INSERT INTO attachments
              (record_id, filename, mime_type, source_path, size_bytes, sha256, content_blob)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (record_id, source.name, MIME_TYPE, str(source), len(content), sha256, content),
        )
        attachment_id = attachment_cursor.lastrowid
        workbook_cursor = connection.execute(
            """
            INSERT INTO spreadsheet_workbooks
              (record_id, attachment_id, filename, report_year, report_month,
               document_type, sheet_count)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                record_id,
                attachment_id,
                source.name,
                year,
                month,
                document_type,
                len(formula_book.sheetnames),
            ),
        )
        workbook_id = workbook_cursor.lastrowid
        populated_count = 0
        formula_count = 0

        for sheet_index, sheet_name in enumerate(formula_book.sheetnames):
            formula_sheet = formula_book[sheet_name]
            value_sheet = value_book[sheet_name]
            merged_ranges = [str(item) for item in formula_sheet.merged_cells.ranges]
            sheet_cursor = connection.execute(
                """
                INSERT INTO spreadsheet_sheets
                  (workbook_id, sheet_index, sheet_name, max_row, max_column,
                   sheet_state, merged_ranges_json)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    workbook_id,
                    sheet_index,
                    sheet_name,
                    formula_sheet.max_row,
                    formula_sheet.max_column,
                    formula_sheet.sheet_state,
                    json.dumps(merged_ranges, ensure_ascii=False),
                ),
            )
            sheet_id = sheet_cursor.lastrowid
            cell_rows = []
            for row in formula_sheet.iter_rows():
                for cell in row:
                    raw_value = cell.value
                    if raw_value is None:
                        continue
                    is_formula = cell.data_type == "f"
                    cached_value = value_sheet[cell.coordinate].value if is_formula else raw_value
                    cell_rows.append(
                        (
                            sheet_id,
                            cell.row,
                            cell.column,
                            cell.coordinate,
                            cell.data_type or type(raw_value).__name__,
                            None if is_formula else text_value(raw_value),
                            text_value(raw_value) if is_formula else None,
                            text_value(cached_value),
                            cell.number_format,
                        )
                    )
                    populated_count += 1
                    formula_count += int(is_formula)
            connection.executemany(
                """
                INSERT INTO spreadsheet_cells
                  (sheet_id, row_number, column_number, cell_address, value_type,
                   value_text, formula_text, cached_value_text, number_format)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                cell_rows,
            )

        connection.execute(
            """
            UPDATE spreadsheet_workbooks
            SET populated_cell_count = ?, formula_count = ?
            WHERE id = ?
            """,
            (populated_count, formula_count, workbook_id),
        )
        connection.execute(
            """
            INSERT INTO change_log (record_id, operation, detail_json)
            VALUES (?, 'INSERT', ?)
            """,
            (
                record_id,
                json.dumps(
                    {
                        "kind": "xlsx_import",
                        "sheet_count": len(formula_book.sheetnames),
                        "populated_cell_count": populated_count,
                        "formula_count": formula_count,
                    },
                    ensure_ascii=False,
                ),
            ),
        )

    formula_book.close()
    value_book.close()
    return (
        f"Imported: {source.name} | sheets={len(formula_book.sheetnames)} "
        f"cells={populated_count} formulas={formula_count}"
    )


def main() -> int:
    sources = [Path(arg).resolve() for arg in sys.argv[1:]]
    if not sources:
        print("Usage: import-xlsx.py <file1.xlsx> [file2.xlsx ...]", file=sys.stderr)
        return 1
    missing = [str(source) for source in sources if not source.is_file()]
    if missing:
        print("Missing files:", *missing, sep="\n", file=sys.stderr)
        return 2

    connection = sqlite3.connect(DB_PATH)
    connection.execute("PRAGMA foreign_keys = ON")
    connection.executescript(SCHEMA)
    try:
        for source in sources:
            print(import_workbook(connection, source))
    finally:
        connection.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
